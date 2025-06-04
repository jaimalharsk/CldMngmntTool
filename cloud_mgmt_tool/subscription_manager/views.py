from collections import defaultdict

from django.contrib.auth import login, get_user_model
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import boto3
from botocore.exceptions import ClientError, NoCredentialsError
import os
from rest_framework.decorators import api_view
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from io import BytesIO
from .forms import CustomUserCreationForm, CloudServiceSubscriptionForm, CloudAccountForm
from .models import CloudServiceSubscription, CloudAccount, CloudAccountUsage
from .serializers import SubscriptionSerializer
from .cloud_api import CloudAPI
from .payment_api import PaymentAPI
from .auth_api import AuthAPI
from django.core.mail import send_mail
#from cloud_mgmt_tool.subscription_manager.gcp_usage import fetch_gcp_usage_and_create_subscriptions
from django.urls import reverse


User = get_user_model()

def send_budget_alert_email(user, subscription):
    subject = f"⚠️ Budget Alert for {subscription.service_name}"
    usage_percent = round((subscription.price / subscription.monthly_budget) * 100, 2)
    message = (
        f"Hi {user.username},\n\n"
        f"Your subscription to {subscription.service_name} on {subscription.cloud_provider.upper()} "
        f"has reached {usage_percent}% of your monthly budget (${subscription.monthly_budget}).\n\n"
        f"Current cost: ${subscription.price}\n"
        f"Alert threshold: {subscription.alert_threshold}%\n\n"
        "Please review your usage or update your budget settings."
    )

    send_mail(
        subject,
        message,
        'noreply@yourcloudtool.com',
        [user.email],
        fail_silently=False,
    )
# ========== Registration ==========
def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


# ========== Custom Login ==========
class AdminStyleLoginView(LoginView):
    template_name = 'admin/login.html'


# ========== Home ==========
def home(request):
    return render(request, 'home.html')


# ========== Dashboard ==========
@login_required
def dashboard(request):
    user = request.user
    CloudServiceSubscription.objects.filter(user=user, cloud_provider='gcp').delete()

    subscriptions = CloudServiceSubscription.objects.filter(user=user).exclude(cloud_provider='gcp')

    if request.method == 'POST':
        if 'add_subscription' in request.POST:
            form = CloudServiceSubscriptionForm(request.POST)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.user = user
                obj.save()
                messages.success(request, "Subscription added!")
                return redirect('dashboard')

        elif 'set_budget' in request.POST:
            sub_id = request.POST.get('subscription_id')
            monthly_budget = request.POST.get('monthly_budget')
            alert_threshold = request.POST.get('alert_threshold')

            try:
                subscription = CloudServiceSubscription.objects.get(id=sub_id, user=user)
                if monthly_budget:
                    subscription.monthly_budget = float(monthly_budget)
                if alert_threshold:
                    subscription.alert_threshold = int(alert_threshold)
                subscription.save()
                messages.success(request, "Budget set successfully!", extra_tags='budget')
            except CloudServiceSubscription.DoesNotExist:
                messages.error(request, "Subscription not found.")
            except ValueError:
                messages.error(request, "Invalid budget or threshold values.")

            return redirect('dashboard')

    services = []

    subscriptions = CloudServiceSubscription.objects.filter(user=user)

    for sub in subscriptions.filter(status='active'):
        budget_amount = sub.monthly_budget or 0
        threshold = sub.alert_threshold or 80
        usage_percent = round((sub.price / budget_amount) * 100, 2) if budget_amount else 0

        if budget_amount:
            if usage_percent >= threshold and not sub.alert_sent:
                send_budget_alert_email(user, sub)
                sub.alert_sent = True
                sub.save()
            # RESET alert_sent if usage goes below threshold AFTER an alert was sent
            elif usage_percent < threshold and sub.alert_sent:
                sub.alert_sent = False
                sub.save()

        services.append({
            'id': sub.id,
            'name': sub.service_name,
            'cost': sub.price,
            'budget': budget_amount,
            'threshold': threshold,
            'usage_percent': usage_percent,
        })

    monthly_cost = sum(sub.price for sub in subscriptions)
    pending_alerts = subscriptions.filter(status='Pending').count()
    budget_limit = sum(sub.monthly_budget for sub in subscriptions)

    usage_chart = {
        'type': 'line',
        'data': {
            'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May'],
            'datasets': [{
                'label': 'Usage (hrs)',
                'data': [10, 14, 8, 12, 16],
                'backgroundColor': 'rgba(54, 162, 235, 0.2)',
                'borderColor': 'rgba(54, 162, 235, 1)',
                'borderWidth': 1,
                'fill': True,
            }]
        }
    }

    billing_chart = {
        'type': 'bar',
        'data': {
            'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May'],
            'datasets': [{
                'label': 'Billing ($)',
                'data': [100, 120, 90, 150, 130],
                'backgroundColor': 'rgba(255, 99, 132, 0.5)',
                'borderColor': 'rgba(255, 99, 132, 1)',
                'borderWidth': 1,
            }]
        }
    }

    return render(request, 'subscription_manager/dashboard.html', {
        'active_subscriptions': subscriptions.filter(status='active'),
        'subscriptions': subscriptions,
        'monthly_cost': monthly_cost,
        'pending_alerts': pending_alerts,
        'budget_limit': budget_limit,
        'usage_chart': usage_chart,
        'billing_chart': billing_chart,
        'services': services,
    })



def connect_cloud(request):
    if request.method == 'POST':
        provider = request.POST.get('provider')

        if provider == 'aws':
            aws_access_key = request.POST.get('aws_access_key')
            aws_secret_key = request.POST.get('aws_secret_key')
            aws_iam_arn = request.POST.get('aws_iam_arn')

            try:
                sts_client = boto3.client(
                    'sts',
                    aws_access_key_id=aws_access_key,
                    aws_secret_access_key=aws_secret_key
                )
                sts_client.get_caller_identity()

                CloudAccount.objects.create(
                    user=request.user,
                    provider=provider,
                    access_key=aws_access_key,
                    secret_key=aws_secret_key,
                    role_arn=aws_iam_arn,
                )
                return JsonResponse({"success": True, "message": "AWS account connected successfully!"})
            except NoCredentialsError:
                return JsonResponse({"success": False, "error": "AWS credentials are missing or invalid."})
            except ClientError as e:
                return JsonResponse({"success": False, "error": f"Failed to connect to AWS: {str(e)}"})

        elif provider == 'gcp':
            project_id = request.POST.get('gcp_project_id')
            if project_id:
                CloudAccount.objects.create(
                    user=request.user,
                    provider=provider,
                    project_id=project_id
                )
                try:
                    dummy_services = [
                        {"service_name": "GCP Compute Engine", "price": 100.00},
                        {"service_name": "GCP Cloud Storage", "price": 50.50},
                        {"service_name": "GCP BigQuery", "price": 80.25},
                    ]
                    for service in dummy_services:
                        CloudServiceSubscription.objects.create(
                            user=request.user,
                            cloud_provider='gcp',
                            service_name=service['service_name'],
                            price=service['price'],
                            status='active',
                            frequency='monthly',
                            monthly_budget=service['price'],
                            alert_threshold=80,
                            is_temporary=True,
                        )
                    return JsonResponse({
                        "success": True,
                        "message": "✅ GCP account connected and 3 dummy subscriptions added.",
                        "redirect_url": reverse('dashboard')
                    })
                except Exception as e:
                    return JsonResponse({
                        "success": False,
                        "error": f"GCP account saved, but adding subscriptions failed: {str(e)}"
                    })

            else:
                return JsonResponse({"success": False, "error": "Invalid GCP project ID."})

        elif provider == 'azure':
            subscription_id = request.POST.get('azure_subscription_id')
            tenant_id = request.POST.get('azure_tenant_id')
            if subscription_id and tenant_id:
                CloudAccount.objects.create(
                    user=request.user,
                    provider=provider,
                    subscription_id=subscription_id,
                    tenant_id=tenant_id
                )
                return JsonResponse({"success": True, "message": "Azure account connected successfully!"})
            else:
                return JsonResponse({"success": False, "error": "Invalid Azure credentials."})

        else:
            return JsonResponse({"success": False, "error": "Unsupported provider."})

    return JsonResponse({"success": False, "error": "Invalid request method."})


# ========== Profile ==========
@login_required
def user_profile(request):
    return render(request, 'subscription_manager/user_profile.html')


# ========== Export Unified Data to XLSX ==========
@login_required
def export_user_data_xlsx(request):
    user = request.user
    subscriptions = CloudServiceSubscription.objects.filter(user=user)

    wb = Workbook()
    ws = wb.active
    ws.title = "User Report"

    # --- USER INFO ---
    ws.append(["User Information"])
    ws.append(["Username", "Email", "Last Login"])
    ws.append([user.username, user.email, str(user.last_login)])
    ws.append([])

    # --- SUBSCRIPTION DATA TABLE ---
    ws.append(["Subscription Report"])
    ws.append([
        "Service Name", "Usage (units)", "Budget (units)",
        "Usage % of Budget", "Cost", "Cost per Unit"
    ])

    row_start = ws.max_row + 1

    for index, sub in enumerate(subscriptions, start=row_start):
        usage = getattr(sub, 'usage', 0) or 0
        budget = sub.monthly_budget or 1  # avoid zero division
        cost = float(sub.price) if sub.price else 0

        usage_percent = round((usage / budget) * 100, 2)
        cost_per_unit = round((cost / usage), 2) if usage else 0

        ws.append([
            sub.service_name,
            usage,
            budget,
            f"{usage_percent}%",
            cost,
            cost_per_unit,
            ""
        ])

        if hasattr(sub, 'logo') and sub.logo and hasattr(sub.logo, 'path'):
            try:
                pil_img = PILImage.open(sub.logo.path)
                pil_img.thumbnail((80, 80))
                img_buffer = BytesIO()
                pil_img.save(img_buffer, format='PNG')
                img_buffer.seek(0)

                img = ExcelImage(img_buffer)
                img_cell = f"G{index}"
                ws.add_image(img, img_cell)
                ws.row_dimensions[index].height = 60
            except Exception as e:
                print(f"Error adding image for {sub.service_name}: {e}")

    screenshot_path = f"media/screenshots/dashboard_{user.username}.png"
    if os.path.exists(screenshot_path):
        try:
            screenshot_img = ExcelImage(screenshot_path)
            ws_ss = wb.create_sheet("Dashboard Screenshot")
            ws_ss.add_image(screenshot_img, "A1")
        except Exception as e:
            print(f"Error adding screenshot: {e}")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{user.username}_cloud_report.xlsx"'
    wb.save(response)
    return response


# ========== API: Get Subscriptions ==========
@api_view(['GET'])
def get_subscriptions(request):
    user = request.user
    subscriptions = CloudServiceSubscription.objects.filter(user=user)
    serializer = SubscriptionSerializer(subscriptions, many=True)
    return Response(serializer.data)


# ========== API: List Cloud Resources ==========
def list_cloud_resources(request, provider):
    api = CloudAPI(provider)
    return JsonResponse({"message": api.list_resources()})


# ========== API: Create Subscription ==========
def create_subscription(request, provider, email):
    payment_api = PaymentAPI(provider)
    return JsonResponse({"message": payment_api.create_subscription(email)})


# ========== Auth Redirect ==========
def login_redirect(request, provider):
    auth_api = AuthAPI(provider)
    return JsonResponse({"message": auth_api.login_url()})

@login_required
def dashboard_ajax(request):
    try:
        cloud_accounts = CloudAccount.objects.filter(user=request.user)
        subscriptions = CloudServiceSubscription.objects.filter(user=request.user)

        usage_totals = defaultdict(float)
        budget_totals = defaultdict(float)

        allowed_gcp_services = {
            "GCP Compute Engine", "GCP Cloud Storage", "GCP BigQuery"
        }

        for sub in subscriptions:
            # Allow only dummy GCP subscriptions, block real/fetched ones
            if sub.cloud_provider == 'gcp' and sub.service_name not in allowed_gcp_services:
                continue

            price = float(sub.price) if sub.price else 0.0
            budget = float(sub.monthly_budget) if sub.monthly_budget else 0.0
            usage_totals[sub.cloud_provider.lower()] += price
            budget_totals[sub.cloud_provider.lower()] += budget

        cloud_data = []
        for account in cloud_accounts:
            provider = account.provider.lower()
            if provider == 'gcp' and not subscriptions.filter(cloud_provider='gcp', service_name__in=allowed_gcp_services).exists():
                continue

            cloud_data.append({
                'account': {'provider': provider},
                'usage': {'total_cost': usage_totals.get(provider, 0)},
                'budget': {'monthly_budget': budget_totals.get(provider, 0)},
            })

        return JsonResponse({'success': True, 'cloud_data': cloud_data})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)




